#!/usr/bin/env python
#coding: utf-8

import os
from time import time
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl import Workbook
from openpyxl import load_workbook
import gevent
from gevent import spawn
from gevent import monkey;monkey.patch_all()
from gevent.pool import Pool
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko.ssh_exception import AuthenticationException
from paramiko.ssh_exception import SSHException


def read_device_excel( ):

    ip_list = []
    
    wb1 = load_workbook('/home/netops/linsy_env/cs_lab.xlsx')
    ws1 = wb1.get_sheet_by_name("Sheet1")
    
    for cow_num in range(2,ws1.max_row+1):

        ipaddr = ws1["a"+str(cow_num)].value
        ip_list.append(ipaddr)
            
    return ip_list


def get_config(ipaddr):
    

    session = ConnectHandler(device_type="huawei",
                            ip=ipaddr,
                            username="rrrrot",
                            password="11111111",
                            banner_timeout=300)

    print("connecting to "+ ipaddr)
    print ("---- Getting HUAWEI configuration from {}-----------".format(ipaddr))

    config_data = session.send_command("dis cu")

    session.disconnect()

    return config_data


def write_config_to_file(config_data,ipaddr):
    now = datetime.now()
    date= "%s-%s-%s"%(now.year,now.month,now.day)
    time_now = "%s-%s"%(now.hour,now.minute)

    #---- Write out configuration information to file
    config_path = '/home/netops/linsy_env/devconfig/' +date
    verify_path = os.path.exists(config_path)
    if not verify_path:
        os.makedirs(config_path)

    config_filename = config_path+"/"+'config_' + ipaddr +"_"+date+"_" + time_now # Important - create unique configuration file name

    print ('---- Writing configuration: ', config_filename)
    with open( config_filename, "w",encoding='utf-8' ) as config_out:  
        config_out.write( config_data )

    return


def write_issue_device(issue_device):
    now = datetime.now()
    date= "%s-%s-%s"%(now.year,now.month,now.day)
    time_now = "%s-%s"%(now.hour,now.minute)

    config_path = '/home/netops/linsy_env/' + "issue_" + date
    verify_path = os.path.exists(config_path)
    if not verify_path:
        os.makedirs(config_path)

    config_filename = config_path+"/"+'issue_'+date+"_" + time_now
    print ('---- Writing issue: ', config_filename)
    with open (config_filename, "w", encoding='utf-8') as issue_facts:
        issue_facts.write('\n'.join(issue_device))


def run_gevent(ipaddr):

    issue_device = []

    try:
        hwconfig = get_config(ipaddr)
        write_config_to_file(hwconfig,ipaddr)

    except (AuthenticationException):
        issue_message = (ipaddr + ': 认证错误 ')
        issue_device.append(issue_message)

    except NetMikoTimeoutException:
        issue_message = (ipaddr + ': 网络不可达 ')
        issue_device.append(issue_message)

    except (SSHException):
        issue_message = (ipaddr +': SSH端口异常 ')
        issue_device.append(issue_message)

    except Exception as unknown_error:
        issue_message = (ipaddr +': 发生未知错误: ')
        issue_device.append(issue_message+str(unknown_error))
    
    finally:
        write_issue_device(issue_device)                  #异常处理信息写入文件

            
def main():

    starting_time = time()   
    ip_list = read_device_excel()
    
    pool = Pool(100)
    pool.map(run_gevent,ip_list)                                 #map(func, iterable)
    pool.join()
    print ('\n---- End get config threading, elapsed time=', time() - starting_time)

#========================================
# Get config of HUAWEI
#========================================
if __name__ == '__main__':
    main()
